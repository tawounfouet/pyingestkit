from __future__ import annotations

import unittest
from datetime import datetime
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook

from pyingestkit.core.exceptions import ConfigurationError, ParseError
from pyingestkit.parsers import ExcelParser


def workbook_bytes(*, title: str = "Data") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(["id", "name", "when", "active"])
    sheet.append([1, " Alice ", datetime(2026, 9, 4, 8, 30), True])
    sheet.append([2, "Bob", None, False])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


class ExcelParserTests(unittest.TestCase):
    def test_optional_dependency_error_is_actionable(self) -> None:
        with patch("pyingestkit.parsers.excel.importlib.import_module", side_effect=ImportError):
            with self.assertRaisesRegex(ConfigurationError, r"pyingestkit\[excel\]"):
                ExcelParser().parse_bytes(b"not-used")

    def test_parses_xlsx_preserving_native_cell_values(self) -> None:
        dataset = ExcelParser().parse_bytes(workbook_bytes())
        self.assertEqual(dataset.fields, ("id", "name", "when", "active"))
        self.assertEqual(dataset.row_count, 2)
        self.assertEqual(dataset[0]["id"], 1)
        self.assertEqual(dataset[0]["name"], " Alice ")
        self.assertIsInstance(dataset[0]["when"], datetime)
        self.assertIs(dataset[0]["active"], True)

    def test_selects_sheet_by_name(self) -> None:
        dataset = ExcelParser(sheet="Data").parse_bytes(workbook_bytes())
        self.assertEqual(dataset.row_count, 2)

    def test_rejects_missing_sheet(self) -> None:
        with self.assertRaises(ParseError):
            ExcelParser(sheet="Missing").parse_bytes(workbook_bytes())

    def test_rejects_duplicate_headers(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["id", "id"])
        sheet.append([1, 2])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        with self.assertRaises(ParseError):
            ExcelParser().parse_bytes(buffer.getvalue())

    def test_header_row_can_be_configured(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["metadata"])
        sheet.append(["id", "name"])
        sheet.append([1, "Alice"])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        dataset = ExcelParser(header_row=2).parse_bytes(buffer.getvalue())
        self.assertEqual(dataset.to_rows(), [{"id": 1, "name": "Alice"}])


if __name__ == "__main__":
    unittest.main()
